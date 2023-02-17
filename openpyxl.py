import openpyxl

# 创建一个Excel workbook 对象
book = openpyxl.Workbook()

# 创建时，会自动产生一个sheet，通过active获取
sh = book.active

# 修改当前 sheet 标题为 工资表
sh.title = '工资表'

# 保存文件
book.save('信息.xlsx')

# 增加一个名为 '年龄表' 的sheet，放在最后
sh1 = book.create_sheet('年龄表-最后')

# 增加一个 sheet，放在最前
sh2 = book.create_sheet('年龄表-最前',0)

# 增加一个 sheet，指定为第2个表单
sh3 = book.create_sheet('年龄表2',1)

# 根据名称获取某个sheet对象
sh = book['工资表']

# 给第一个单元格写入内容
sh['A1'] = '你好'

# 获取某个单元格内容
print(sh['A1'].value)

# 根据行号列号， 给第一个单元格写入内容，
# 注意和 xlrd 不同，是从 1 开始
sh.cell(2,2).value = '白月黑羽'

# 根据行号列号， 获取某个单元格内容
print(sh.cell(1, 1).value)

book.save('信息.xlsx')

插入行、列
sheet 对象的 insert_rows 和 insert_cols 方法，分别用来插入行和列

删除行、列
sheet 对象的 delete_rows 和 delete_cols 方法，分别用来删除行和列

文字颜色、字体、大小
单元格里面的样式风格（包括颜色、字体、大小、下划线等）都是通过Font对象设定的

import openpyxl
# 导入Font对象 和 colors 颜色常量
from openpyxl.styles import Font,colors

wb = openpyxl.load_workbook('income.xlsx')
sheet = wb['2018']

# 指定单元格字体颜色，
sheet['A1'].font = Font(color=colors.RED, #使用预置的颜色常量
                        size=15,    # 设定文字大小
                        bold=True,  # 设定为粗体
                        italic=True # 设定为斜体
                        )

# 也可以使用RGB数字表示的颜色
sheet['B1'].font = Font(color="981818")

# 指定整行 字体风格， 这里指定的是第3行
font = Font(color="981818")
for y in range(1, 100): # 第 1 到 100 列
    sheet.cell(row=3, column=y).font = font

# 指定整列 字体风格， 这里指定的是第2列
font = Font(bold=True)
for x in range(1, 100): # 第 1 到 100 行
    sheet.cell(row=x, column=2).font = font

# 指定 某个单元格背景色
sheet['A1'].fill = PatternFill("solid", "E39191")

# 指定 整行 背景色， 这里指定的是第2行
fill = PatternFill("solid", "E39191")
for y in range(1, 100): # 第 1 到 100 列
    sheet.cell(row=2, column=y).fill = fill

# 在第1行，第4列 的位置插入图片
sheet.add_image(Image('1.png'), 'D1')

wb.save('income-1.xlsx')
